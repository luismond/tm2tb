"""Reads parallel bilingual data files."""
import os
from collections import OrderedDict
import xmltodict
import openpyxl
import pandas as pd
from pandas.errors import ParserError


class BitextReader:
    """Read parallel bilingual data files."""

    extensions = ['.csv', '.mqxliff', '.mxliff', '.tmx', '.xlsx']

    def __init__(self, file_path):
        self.file_path = file_path
        self.file_extension = os.path.splitext(self.file_path)[1]
        self.file_size = os.path.getsize(self.file_path)
        self.file_max_size = 100000000

    def read_bitext(self):
        """Get bitext from file content."""
        bitext = self.get_file_content()
        bitext.columns = ['src', 'trg']
        bitext = bitext.fillna('')
        if len(bitext) == 0:
            raise ValueError('The document appears to be empty.')
        bitext = bitext.astype(str)
        bitext = list(zip(bitext['src'], bitext['trg']))
        return bitext

    def get_file_content(self):
        """Check extension and get bitext."""
        if self.file_size > self.file_max_size:
            max_mb = self.file_max_size/1000000
            raise ValueError(f"Oversized file. Max size: {max_mb} MB.")
        if self.file_extension not in self.extensions:
            raise ValueError('Unsupported file extension.')
        if self.file_extension == '.csv':
            content = CsvReader(self.file_path).read_csv()
        if self.file_extension == '.mqxliff':
            content = XmlReader(self.file_path).read_mqxliff()
        if self.file_extension == '.mxliff':
            content = XmlReader(self.file_path).read_mxliff()
        if self.file_extension == '.tmx':
            content = XmlReader(self.file_path).read_tmx()
        if self.file_extension == '.xlsx':
            content = XlsxReader(self.file_path).read_xlsx()
        return content


class CsvReader:
    """Read .csv files."""

    def __init__(self, file_path):
        self.file_path = file_path

    def read_csv(self):
        """Read two-column .csv file."""
        try:
            content = pd.read_csv(self.file_path, encoding='utf-8')
        except UnicodeError:
            content = pd.read_csv(self.file_path, encoding='uft-16')
        except ParserError:
            content = pd.read_csv(self.file_path, sep='\t')

        content = self.validate_csv_content(content)
        return content

    @staticmethod
    def validate_csv_content(content):
        """Validate number of rows and columns."""
        content = content.dropna()
        n_cols = len(content.columns)
        n_rows = len(content)
        if n_rows == 0:
            raise ValueError("The .csv file appears to be empty.")
        if n_cols != 2:
            msg = '.csv file has {} cols. Max. columns: 2.'
            raise ValueError(msg.format(n_cols))
        return content


class XmlReader:
    """Read an xml-based bilingual file (.mqxliff, .mxliff, .tmx)."""

    def __init__(self, file_path):
        self.file_path = file_path

    def read_file(self):
        """Attempt to read file content."""
        try:
            with open(self.file_path, encoding='utf-8') as file:
                file_content = file.read()
        except UnicodeError:
            with open(self.file_path, encoding='utf-16') as file:
                file_content = file.read()
        return file_content

    @staticmethod
    def parse_xml(file_content):
        """Parse XML file content."""
        try:
            xml = xmltodict.parse(file_content, disable_entities=True)
        except xmltodict.expat.ExpatError:
            raise ValueError('Expat Error')
        return xml

    def read_mqxliff(self):
        """Read bilingual .mqxliff file."""
        file_content = self.read_file()
        xml = self.parse_xml(file_content)
        units = xml['xliff']['file']['body']['trans-unit']
        try:
            source_segments = [unit['source'] for unit in units]
            target_segments = [unit['target'] for unit in units]
        except:  # if the file has only one segment
            source_segments = [units['source']]
            target_segments = [units['target']]
        segments = zip(source_segments, target_segments)
        segments = [(seg[0]['#text'], seg[1]['#text']) for seg in segments
                    if '#text' in seg[0] and '#text' in seg[1]]
        content = pd.DataFrame(segments)
        if len(content) == 0:
            if all("#text" in seg for seg in source_segments) is False:
                raise ValueError('The source segments are empty.')
            if all("#text" in seg for seg in target_segments) is False:
                raise ValueError('The target segments are empty.')
        return content

    def read_mxliff(self):
        """Read bilingual .mxliff file."""
        file_content = self.read_file()
        xml = self.parse_xml(file_content)
        file = xml['xliff']['file']
        if isinstance(file, dict) is True:
            units = [unit['trans-unit'] for unit in file['body']['group']]
        if isinstance(file, list) is True:  # in case it is a joined file
            units = []
            for sub_file in file:
                for unit in sub_file['body']['group']:
                    units.append(unit['trans-unit'])
        source_segments = [unit['source'] for unit in units]
        target_segments = [unit['target'] for unit in units]
        segments = zip(source_segments, target_segments)
        content = pd.DataFrame(segments)
        return content

    def read_tmx(self):
        """Read bilingual .tmx file."""
        file_content = self.read_file()
        xml = self.parse_xml(file_content)
        source_segments = [t['tuv'][0]['seg'] for t in xml['tmx']['body']['tu']]
        target_segments = [t['tuv'][1]['seg'] for t in xml['tmx']['body']['tu']]
        source_segments = [self.get_tmx_text(unit) for unit in source_segments]
        target_segments = [self.get_tmx_text(unit) for unit in target_segments]
        segments = zip(source_segments, target_segments)
        content = pd.DataFrame(segments)
        return content

    @staticmethod
    def get_tmx_text(unit):
        """
        Retrieve string from translation unit.

        unit : string or OrderedDict
            Translation unit in tmx file. Can be a string or OrderedDict.

        Returns
        -------
        unit_text : string from translation unit or empty string
        """
        try:
            if isinstance(unit, str) is True:
                return unit
            if isinstance(unit, OrderedDict) is True:
                return unit['#text']
            return None
        except:
            return ''


class XlsxReader:
    """Read .xlsx file."""

    def __init__(self, file_path):
        self.file_path = file_path
        self.max_col = 2      # Max col limit to avoid OOM errors
        self.max_row = 10000  # Max row limit to avoid OOM errors

    def get_workbook(self):
        """Load .xlsx workbook."""
        try:
            return openpyxl.load_workbook(self.file_path)
        except:
            raise ValueError("""Unable to load xlsx file.""")

    def get_sheet(self):
        """Get first workbook sheet."""
        try:
            workbook = self.get_workbook()
            return workbook[workbook.sheetnames[0]]
        except:
            raise ValueError("""Unable to read workbook sheet.""")

    def validate_sheet(self, sheet):
        """Validate number of rows and columns."""
        # Get all cells from all rows
        if sheet.max_column > self.max_col:
            max_col_e = ".xlsx file has {} columns. It must have 2 columns."
            raise ValueError(max_col_e.format(sheet.max_column))
        if sheet.max_row > self.max_row:
            max_row_e = ".xlsx file has {} rows. The limit is 10000 rows."
            raise ValueError(max_row_e.format(sheet.max_row))
        return sheet

    def read_xlsx(self):
        """Read xlsx file."""
        sheet = self.validate_sheet(self.get_sheet())
        rows = list(sheet.iter_rows())
        content = pd.DataFrame([[cell.value for cell in row] for row in rows])
        return content
