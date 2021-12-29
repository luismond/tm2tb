#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BilingualReader: Reads parallel bilingual data files
"""
import os
from collections import OrderedDict
import xmltodict
import openpyxl
import pandas as pd

class BilingualReader:
    'Reads parallel bilingual data files'
    extensions = ['.csv', '.mqxliff', '.mxliff', '.tmx', '.xlsx']
    def __init__(self, path, fn):
        self.file_path = os.path.join(path, fn)
        self.file_extension = os.path.splitext(self.file_path)[1]
        self.file_size = os.path.getsize(self.file_path)
        self.file_max_size = 5000000

    def get_file_content(self):
        'Check extension and get bitext'
        if self.file_size > self.file_max_size:
            raise ValueError('Oversized file. Max size: {}'.format(self.file_max_size))
        if not self.file_extension in self.extensions:
            raise ValueError('Unsupported file extension')
        if self.file_extension == '.csv':
            csvr = CsvReader(self.file_path, self.file_extension)
            content = csvr.read_csv()
        if self.file_extension == '.mqxliff':
            mqr = MqxliffReader(self.file_path, self.file_extension)
            content = mqr.read_mqxliff()
        if self.file_extension == '.mxliff':
            mxr = MxliffReader(self.file_path, self.file_extension)
            content = mxr.read_mxliff()
        if self.file_extension == '.tmx':
            tmxr = TmxReader(self.file_path, self.file_extension)
            content = tmxr.read_tmx()
        if self.file_extension == '.xlsx':
            xlsxr = XlsxReader(self.file_path, self.file_extension)
            content = xlsxr.read_xlsx()
        return content

    def get_bitext(self):
        'Get bitext from file content'
        bitext = self.get_file_content()
        bitext.columns = ['src', 'trg']
        bitext = bitext.dropna()
        if len(bitext)==0:
            raise ValueError('Document appears to be empty')
        bitext = bitext.astype(str)
        bitext = bitext[bitext['src'].str.len()>0]
        bitext = bitext[bitext['trg'].str.len()>0]
        return bitext

class CsvReader:
    'Reads .csv files'
    def __init__(self, file_path, file_extension):
        self.file_path = file_path
        self.file_extension = file_extension

    def read_csv(self):
        'Read two-column .csv file'
        try:
            content = pd.read_csv(self.file_path, encoding='utf-8')
        except UnicodeError:
            content = pd.read_csv(self.file_path, encoding='uft-16')
        content = self.validate_csv_content(content)
        return content

    def validate_csv_content(self, content):
        'Validate nrows and ncols'
        content = content.dropna()
        n_cols = len(content.columns)
        n_rows = len(content)
        if n_rows == 0:
            raise ValueError("The .csv file appears to be empty")
        if n_cols != 2:
            msg = '.csv file has {} cols. Max. columns: 2'
            raise ValueError(msg.format(n_cols))
        return content

def parse_xml(file_path):
    'Parse xml-based bilingual file (.mqxliff, .mxliff, .tmx)'
    try:
        with open(file_path, encoding='utf-8') as file:
            doc = xmltodict.parse(file.read())
    except UnicodeError:
        with open(file_path, encoding='utf-16') as file:
            doc = xmltodict.parse(file.read())
    except xmltodict.expat.ExpatError:
        raise ValueError('xmltodict Expat Error')
    return doc

class MqxliffReader:
    'Read bilingual .mqxliff file'
    def __init__(self, file_path, file_extension):
        self.file_path = file_path
        self.file_extension = file_extension

    def split_format_tags(self):
        'Temporary hack to avoid joined words due to format tags'
        with open(self.file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            content = content.replace('<ph', ' <ph')
            content = content.replace('</ph>', ' </ph> ')
        with open(self.file_path, 'w', encoding='utf-8') as file:
            file.write(content)

    def read_mqxliff(self):
        'Read bilingual .mqxliff file'
        self.split_format_tags()
        doc = parse_xml(self.file_path)
        units = doc['xliff']['file']['body']['trans-unit']
        try:
            source_segments = [unit['source'] for unit in units]
            target_segments = [unit['target'] for unit in units]
        except:
            # in case file has only one segment
            source_segments = [units['source']]
            target_segments = [units['target']]
        segments = zip(source_segments, target_segments)
        segments = [(seg[0]['#text'], seg[1]['#text']) for seg in segments
                    if '#text' in seg[0] and '#text' in seg[1]]
        bitext = pd.DataFrame(segments)
        return bitext

class MxliffReader:
    'Read bilingual .mxliff file'
    def __init__(self, file_path, file_extension):
        self.file_path = file_path
        self.file_extension = file_extension

    def read_mxliff(self):
        'Read bilingual .mxliff file'
        doc = parse_xml(self.file_path)
        units = [unit['trans-unit'] for unit in doc['xliff']['file']['body']['group']]
        source_segments = [unit['source'] for unit in units]
        target_segments = [unit['target'] for unit in units]
        segments = zip(source_segments, target_segments)
        bitext = pd.DataFrame(segments)
        return bitext

class TmxReader:
    'Reads bilingual .tmx file'
    def __init__(self, file_path, file_extension):
        self.file_path = file_path
        self.file_extension = file_extension

    def read_tmx(self):
        'Reads bilingual .tmx file'
        doc = parse_xml(self.file_path)
        source_segments = [t['tuv'][0]['seg'] for t in doc['tmx']['body']['tu']]
        target_segments = [t['tuv'][1]['seg'] for t in doc['tmx']['body']['tu']]
        source_segments = [self.get_tmx_text(unit) for unit in source_segments]
        target_segments = [self.get_tmx_text(unit) for unit in target_segments]
        segments = zip(source_segments, target_segments)
        bitext = pd.DataFrame(segments)
        return bitext

    def get_tmx_text(self, unit):
        """
        unit : string or OrderedDict
            Translation unit in tmx file. Can be a string or OrderedDict.
        Returns
        -------
        unit_text : string from translation unit or empty string
        """
        try:
            if isinstance(unit, str) is True:
                unit_text = unit
                return unit_text
            if isinstance(unit, OrderedDict) is True:
                unit_text = unit['#text']
                return unit_text
            return None
        except:
            return ''

class XlsxReader:
    'Reads .xlsx file'
    def __init__(self, file_path, file_extension):
        self.file_path = file_path
        self.file_extension = file_extension
        self.max_col = 2    # Max col limit to avoid OOM errors
        self.max_row = 10000 # Max row limit to avoid OOM errors

    def get_workbook(self):
        'Loads .xlsx workbook'
        try:
            return openpyxl.load_workbook(self.file_path)
        except:
            raise ValueError('Unable to load xlsx file')

    def get_sheet(self):
        'Gets first workbook sheet'
        try:
            workbook = self.get_workbook()
            return workbook[workbook.sheetnames[0]]
        except:
            raise ValueError('Unable to read workbook sheet')

    def validate_sheet(self, sheet):
        'Validates nrows and ncolumns'
        # Get all cells from all rows
        if sheet.max_column > self.max_col:
            max_col_e = '.xlsx file has {} columns. It must have 2 columns.'
            raise ValueError(max_col_e.format(sheet.max_column))
        if sheet.max_row > self.max_row:
            max_row_e = '.xlsx file has {} rows. The limit is 10000 rows.'
            raise ValueError(max_row_e.format(sheet.max_row))
        return sheet

    def read_xlsx(self):
        'Reads xlsx file'
        sheet = self.validate_sheet(self.get_sheet())
        rows = list(sheet.iter_rows())
        content = pd.DataFrame([[cell.value for cell in row] for row in rows])
        return content
